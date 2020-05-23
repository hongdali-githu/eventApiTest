#coding=utf-8
import openpyxl
import xlrd
from event.config.config import LoggerConfig,TestData
from event.common.LoggerHandle import Logger_Handle
logger=Logger_Handle("root","INFO",filename=LoggerConfig.log_path)
#处理xlsx结尾的execl
class Execl_Handle_xlsx:
    #file:execl路径
    #sheetnum:execl标签页,从0开始
    def __init__(self,file,sheetnum):
        self.file=file
        self.sheetnum=sheetnum
    #打开execl，获取页对象
    def open_execl(self):
        try:
            workbook=openpyxl.load_workbook(self.file)
            sheet=workbook.worksheets[self.sheetnum]
        except Exception as e:
            logger.error("文件打开失败{}".format(e))
            raise e
        else:
            logger.info("文件打开成功")
            return sheet
    #获取表头
    def get_sheetHead(self):
        head_data=[]
        try:
            sheethead=self.open_execl()[1]
            for cell in sheethead:
                head_data.append(cell.value)
            return head_data
        except Exception as e:
            logger.error("获取表头失败{}".format(e))
            raise e

    #获取execl所有内容，没一行以字典形式储存
    def get_all(self):
        all_data=[]
        #从第二行开始获取所有行
        if self.open_execl().max_row==2:
            row_data = []
            try:
                for cell in self.open_execl()[2]:
                    row_data.append(cell.value)
                row_data = zip(self.get_sheetHead(), row_data)
                row_data = dict(row_data)
                all_data.append(row_data)
                return all_data
            except Exception as e:
                logger.exception("获取execl所有内容失败{}".format(e))
                raise e

        else:
            for row in self.open_execl()[2:self.open_execl().max_row]:
                row_data=[]
                try:
                    for cell in row:
                        #获取单行内容
                        row_data.append(cell.value)
                    #将表头与表的内容转行成字典
                    row_data=zip(self.get_sheetHead(),row_data)
                    row_data=dict(row_data)
                    all_data.append(row_data)
                    return all_data
                except Exception as e:
                    logger.exception("获取execl全部内容失败{}".format(e))
                    raise e



#处理xls结尾的execl
class Execl_Handle_xls:
    #file:execl路径
    #sheetnum:execl标签页,从0开始
    def __init__(self,file,sheetnum):
        self.file=file
        self.sheetnum=sheetnum

    # 打开execl，获取页签对象
    def open_execl(self):
        workbook = xlrd.open_workbook(self.file)
        sheet = workbook.sheets()[self.sheetnum]
        return sheet

        # 获取表头
    def get_sheetHead(self):
        head_data = []
        for cell in self.open_execl().row(0):
            head_data.append(cell.value)
        return head_data

    #获取execl所有内容，没一行以字典形式储存
    def get_all(self):
        all_data=[]
        #从第二行开始获取所有行
        for row_num in range(1,self.open_execl().nrows):
            row_data=[]
            for cell in self.open_execl().row(row_num):
                #获取单行内容
                row_data.append(cell.value)
            #将表头与表的内容转行成字典
            row_data=zip(self.get_sheetHead(),row_data)
            row_data=dict(row_data)
            all_data.append(row_data)
        return all_data

if __name__=="__main__":
    execl = Execl_Handle_xlsx(TestData.event_task_controller, 0)
    print(execl.get_all())








